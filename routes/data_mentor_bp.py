from flask import Blueprint, send_file, make_response, request, jsonify, render_template, current_app, Response # Blueprint para modularizar y relacionar con app
from flask_bcrypt import Bcrypt                                  # Bcrypt para encriptación
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity   # Jwt para tokens
from database import db                                          # importa la db desde database.py
from logging_config import logger
import os                                                        # Para datos .env
from dotenv import load_dotenv                                   # Para datos .env
load_dotenv()
from utils.data_mentor_utils import query_assistant_mentor
import urllib.request
import urllib.error
import json
import pandas as pd
from models import Usuarios_Por_Asignacion, Usuarios_Sin_ID, ValidaUsuarios,DetalleApies, AvanceCursada, DetallesDeCursos, CursadasAgrupadas,FormularioGestor,CuartoSurveySql, QuintoSurveySql, Comentarios2023, Comentarios2024, Comentarios2025, BaseLoopEstaciones, FichasGoogleCompetencia, FichasGoogle, SalesForce, ComentariosCompetencia, FileDailyID
import hashlib
from sqlalchemy.exc import SQLAlchemyError
import csv
import time
import tempfile
from openai import OpenAI
import httpx
from tempfile import NamedTemporaryFile
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta 
from concurrent.futures import ThreadPoolExecutor
import shutil


OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("Debes definir la variable de entorno OPENAI_API_KEY con tu clave de API.")

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {OPENAI_API_KEY}",
    "OpenAI-Beta": "assistants=v2"
}


data_mentor_bp = Blueprint('data_mentor_bp', __name__)     # instanciar admin_bp desde clase Blueprint para crear las rutas.
bcrypt = Bcrypt()
jwt = JWTManager()

# Sistema de key base pre rutas ------------------------:

API_KEY = os.getenv('API_KEY')

def check_api_key(api_key):
    return api_key == API_KEY

@data_mentor_bp.before_request
def authorize():
    if request.method == 'OPTIONS':
        return
    if request.path in ['/horas-por-curso','/test_data_mentor_bp']:
        return
    api_key = request.headers.get('Authorization')
    if not api_key or not check_api_key(api_key):
        return jsonify({'message': 'Unauthorized'}), 401
    
# RUTA TEST:

@data_mentor_bp.route('/test_data_mentor_bp', methods=['GET'])
def test():
    logger.info("Chat data mentor bp rutas funcionando ok segun test.")
    return jsonify({'message': 'test bien sucedido','status':"Si lees esto, chat data mentor rutas funcionan bien..."}),200

@data_mentor_bp.route("/chat_mentor", methods=["POST"])
def chat_mentor():
    logger.info("1 - Entró en la ruta Chat_mentor.")
    """
    Recibe prompt y opcionalmente thread_id.
    """
    data = request.get_json()
    if not data or "prompt" not in data:
        return jsonify({"error": "Falta el prompt en el cuerpo de la solicitud"}), 400

    prompt = data["prompt"]
    thread_id = data.get("thread_id")  # puede ser None
    logger.info("2 - Encontró la data del prompt...")
    try:
        response_text, current_thread = query_assistant_mentor(prompt, thread_id)
        return jsonify({"response": response_text, "thread_id": current_thread}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@data_mentor_bp.route("/close_chat_mentor", methods=["POST"])
def close_chat():
    """
    Ruta para cerrar el thread del chat.
    Se espera recibir un JSON con la clave "thread_id".
    Llama al endpoint DELETE de la API para cerrar el hilo usando urllib.
    """
    data = request.get_json()
    if not data or "thread_id" not in data:
        return jsonify({"error": "Falta el thread_id en el cuerpo de la solicitud"}), 400

    thread_id = data["thread_id"]
    delete_url = f"https://api.openai.com/v1/threads/{thread_id}"

    try:
        req = urllib.request.Request(delete_url, headers=HEADERS, method="DELETE")
        with urllib.request.urlopen(req) as response:
            result_data = response.read().decode("utf-8")
            result = json.loads(result_data)
        return jsonify(result), 200
    except urllib.error.HTTPError as e:
        error_message = e.read().decode("utf-8")
        return jsonify({"error": f"HTTPError {e.code}: {error_message}"}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@data_mentor_bp.route('/horas-por-curso', methods=['GET'])
def horas_por_curso():
    data = [
        {"curso": "Node.js Básico", "horas": 5},
        {"curso": "React Intermedio", "horas": 7},
        {"curso": "Flask Fullstack", "horas": 9}
    ]
    return jsonify(data)

# -------------------------- MODELOS QUE SE TIENEN EN CUENTA PARA CONTABILIZAR SUS REGISTROS ----------------------------------

# Diccionario para mapear nombres de string a clases reales
MODELS = {
    'Usuarios_Por_Asignacion': Usuarios_Por_Asignacion,
    'Usuarios_Sin_ID': Usuarios_Sin_ID,
    'ValidaUsuarios': ValidaUsuarios,
    'DetalleApies' : DetalleApies,
    'AvanceCursada': AvanceCursada,
    'DetallesDeCursos' : DetallesDeCursos,
    'CursadasAgrupadas' : CursadasAgrupadas,
    'FormularioGestor' :FormularioGestor,
    'CuartoSurveySql': CuartoSurveySql,
    'QuintoSurveySql' : QuintoSurveySql,
    'Comentarios2023': Comentarios2023,
    'Comentarios2024': Comentarios2024,
    'Comentarios2025': Comentarios2025,
    'BaseLoopEstaciones' : BaseLoopEstaciones,
    'FichasGoogleCompetencia' : FichasGoogleCompetencia,
    'FichasGoogle' : FichasGoogle,
    'SalesForce' : SalesForce,
    'ComentariosCompetencia' : ComentariosCompetencia
    # Agregá los modelos que quieras habilitar acá
}

# -------------------------- Contabilizar longitud de cualquier tabla ----------------------------------

@data_mentor_bp.route('/contar-registros', methods=['POST'])
def contar_registros():
    data = request.get_json()
    nombre_tabla = data.get('tabla')

    if not nombre_tabla:
        return jsonify({"error": "Falta el nombre de la tabla"}), 400

    modelo = MODELS.get(nombre_tabla)
    if not modelo:
        return jsonify({"error": f"La tabla '{nombre_tabla}' no está habilitada o no existe"}), 404

    try:
        cantidad = db.session.query(modelo).count()
        return jsonify({"tabla": nombre_tabla, "cantidad_registros": cantidad}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# -------------------------- ACA VIENEN LAS RUTAS DE LAS TABLAS DE REPORTES ----------------------------


@data_mentor_bp.route('/usuarios_por_asignacion/<int:registro_id>', methods=['GET'])
def get_usuario_por_asignacion(registro_id):
    """
    Devuelve el registro de Usuarios_Por_Asignacion con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Por_Asignacion id={registro_id}")
    registro = Usuarios_Por_Asignacion.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Por_Asignacion id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200

@data_mentor_bp.route('/usuarios_sin_id/<int:registro_id>', methods=['GET'])
def get_usuario_sin_id(registro_id):
    """
    Devuelve el registro de Usuarios_Sin_ID con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Sin_ID id={registro_id}")
    registro = Usuarios_Sin_ID.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Sin_ID id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200




# RUTAS PARA CARGAR TABLAS DE EXPERIENCIA 2023 24 y 25

@data_mentor_bp.route('/cargar_comentarios_2023', methods=['POST'])
def cargar_comentarios_encuesta_2023():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB
    """
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        registros = []
        for _, fila in df.iterrows():
            fecha_raw = fila.get('FECHA')
            try:
                fecha = pd.to_datetime(fecha_raw) if pd.notnull(fecha_raw) else None
            except:
                fecha = None

            nuevo = Comentarios2023(
                fecha=fecha,
                apies=str(fila.get('APIES', '')).strip(),
                comentario=str(fila.get('COMENTARIO', '')).strip(),
                canal=str(fila.get('CANAL', '')).strip(),
                topico=str(fila.get('TÓPICO', '')).strip(),
                sentiment=str(fila.get('SENTIMENT', '')).strip()
            )
            registros.append(nuevo)

        db.session.add_all(registros)
        db.session.commit()

        return jsonify({'mensaje': f'Se guardaron {len(registros)} comentarios', 'status': 200}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_2024', methods=['POST'])
def cargar_comentarios_encuesta_2024():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Optimizado para bajo uso de memoria mediante procesamiento fila a fila con openpyxl.
    Incluye logs de progreso consolidados y tiempo de ejecución en la respuesta.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    total_registros_guardados = 0
    BATCH_SIZE = 5000 
    
    # Contadores y muestras para errores de fecha
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 # Cuántos ejemplos de errores de fecha guardar

    logger.info("======================================================")
    logger.info("============= INICIANDO PROCESO DE CARGA ==============")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        if col_fecha is None or col_apies is None or col_comentario is None or \
           col_canal is None or col_topico is None or col_sentiment is None:
            missing_cols = []
            if col_fecha is None: missing_cols.append('FECHA')
            if col_apies is None: missing_cols.append('APIES')
            if col_comentario is None: missing_cols.append('COMENTARIO')
            if col_canal is None: missing_cols.append('CANAL')
            if col_topico is None: missing_cols.append('TÓPICO')
            if col_sentiment is None: missing_cols.append('SENTIMENT')
            
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_chunk = []
        row_counter_total = 0 
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila
        # enumerate(..., start=2) para obtener el número de fila real del Excel (incluyendo encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            row_counter_total += 1
            
            fecha_raw = row[col_fecha].value if col_fecha is not None else None
            apies = row[col_apies].value if col_apies is not None else ''
            comentario = row[col_comentario].value if col_comentario is not None else ''
            canal = row[col_canal].value if col_canal is not None else ''
            topico = row[col_topico].value if col_topico is not None else ''
            sentiment = row[col_sentiment].value if col_sentiment is not None else ''
            
            fecha = None # Inicializar fecha como None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw # openpyxl ya lo parseó a datetime
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip(): # Si tiene valor no nulo y no es string vacío
                    # Intentar parsear si no es datetime
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                # Si pd.to_datetime con errors='coerce' falla, fecha será NaT, que se convierte en None al asignar
            except Exception as date_e:
                # Solo logear si es un error que no sea solo 'coerce' a NaT, o si queremos más detalle
                pass # No loguear cada error de fecha individualmente
            
            # Contar errores de fecha y guardar una muestra
            if fecha is pd.NaT: # pd.NaT indica que pd.to_datetime falló y coecionó
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                fecha = None # Asegurarse de que sea None para la DB si falló el parseo

            nuevo = Comentarios2024(
                fecha=fecha,
                apies=str(apies).strip(),
                comentario=str(comentario).strip(),
                canal=str(canal).strip(),
                topico=str(topico).strip(),
                sentiment=str(sentiment).strip()
            )
            registros_chunk.append(nuevo)

            # Si el chunk alcanza el tamaño definido, lo guardamos en la DB
            if len(registros_chunk) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando CHUNK {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
                
                db.session.add_all(registros_chunk)
                db.session.commit()
                total_registros_guardados += len(registros_chunk)
                logger.info(f"CHUNK {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB. Total acumulado: {total_registros_guardados}")
                
                db.session.expunge_all() 
                del registros_chunk 
                registros_chunk = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del CHUNK {chunk_counter} liberada. ---")

        # Guardar los registros restantes (el último lote)
        if registros_chunk:
            chunk_counter += 1
            logger.info(f"--- Procesando CHUNK FINAL {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
            db.session.add_all(registros_chunk)
            db.session.commit()
            total_registros_guardados += len(registros_chunk)
            logger.info(f"CHUNK FINAL {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB.")
            logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
            db.session.expunge_all()
            del registros_chunk

        end_time = datetime.now() # Finalizar el temporizador
        time_elapsed = end_time - start_time
        # Formatear el tiempo de ejecución a HH:MM:SS
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("============ PROCESO DE CARGA FINALIZADO ============")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS PROCESADAS EN EXCEL: {row_counter_total}")
        logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': row_counter_total,
                'total_registros_db_guardados': total_registros_guardados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() # Capturar tiempo de finalización incluso en error
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error
        }), 500
    
@data_mentor_bp.route('/cargar_comentarios_2025', methods=['POST'])
def cargar_comentarios_encuesta_2025():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Implementa lógica incremental (deduplicación por hash_id) y optimización de memoria
    mediante openpyxl y procesamiento por lotes.
    Incluye logs detallados de progreso y tiempo de ejecución.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    # Contadores para el resumen final
    total_registros_guardados = 0
    total_filas_excel_leidas = 0
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 
    duplicados_en_archivo_contados = 0
    duplicados_en_db_contados = 0

    # Este set rastreará todos los hashes únicos encontrados en el ARCHIVO ACTUAL
    # para evitar duplicados que aparezcan en diferentes lotes del mismo archivo.
    all_hashes_seen_in_current_file_session = set() 

    BATCH_SIZE = 5000 

    logger.info("======================================================")
    logger.info("====== INICIANDO PROCESO DE CARGA COMENTARIOS 2025 ======")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        # Mapeo de columnas, asegurando que existan
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        # Verificar que las columnas obligatorias existan
        required_cols = {'FECHA', 'APIES', 'COMENTARIO', 'CANAL', 'TÓPICO', 'SENTIMENT'}
        missing_cols = [col for col in required_cols if header_map.get(col) is None]
        if missing_cols:
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_para_lote_db = [] # Esta lista contendrá los objetos ORM para el bulk_save_objects
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila (después del encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            total_filas_excel_leidas += 1
            
            # Extraer valores de celda de forma segura
            fecha_raw = row[col_fecha].value
            apies = row[col_apies].value
            comentario = row[col_comentario].value
            canal = row[col_canal].value
            topico = row[col_topico].value
            sentiment = row[col_sentiment].value
            
            # --- Manejo de la fecha y generación del hash ---
            fecha = None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw 
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip():
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                
                if fecha is pd.NaT: # Si el parseo de Pandas falló y coecionó
                    date_errors_count += 1
                    if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                        sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                    fecha = None # Asegurarse de que sea None para la DB si falló el parseo
            except Exception as date_e:
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}' - {date_e}")
                fecha = None 
            
            # Stripear strings antes de generar hash y crear objeto
            apies = str(apies if apies is not None else '').strip()
            comentario = str(comentario if comentario is not None else '').strip()
            canal = str(canal if canal is not None else '').strip()
            topico = str(topico if topico is not None else '').strip()
            sentiment = str(sentiment if sentiment is not None else '').strip()

            # Generar hash único para este registro
            hash_id = Comentarios2025.generar_hash(fecha, apies, comentario, canal, topico, sentiment)

            # Verificar si este hash_id ya fue visto en el archivo ACTUAL (en cualquier lote procesado)
            if hash_id in all_hashes_seen_in_current_file_session:
                duplicados_en_archivo_contados += 1
                # logger.debug(f"Saltando duplicado DENTRO DEL ARCHIVO (hash: {hash_id}) en fila {row_index_excel}.")
                continue # Saltar esta fila, ya fue procesada o es un duplicado interno

            all_hashes_seen_in_current_file_session.add(hash_id) # Registrar este hash_id como visto

            comentario_obj = Comentarios2025(
                fecha=fecha,
                apies=apies,
                comentario=comentario,
                canal=canal,
                topico=topico,
                sentiment=sentiment,
                hash_id=hash_id # Asignar el hash_id al objeto
            )
            registros_para_lote_db.append(comentario_obj)

            # Si el lote para DB alcanza el tamaño definido, procesamos y guardamos
            if len(registros_para_lote_db) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando LOTE {chunk_counter} ({len(registros_para_lote_db)} candidatos para DB) ---")
                
                # Obtener los hashes de este lote para consultar la DB
                hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
                
                # Consultar la DB para ver cuáles de estos hashes ya existen
                existentes_en_db_lote = set(
                    r[0] for r in db.session.query(Comentarios2025.hash_id)
                    .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                    .all()
                )
                
                # Filtrar los que realmente son nuevos para insertar
                nuevos_para_insertar = [
                    obj for obj in registros_para_lote_db 
                    if obj.hash_id not in existentes_en_db_lote
                ]
                
                duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

                if nuevos_para_insertar:
                    # Usamos bulk_save_objects para una inserción más eficiente
                    db.session.bulk_save_objects(nuevos_para_insertar)
                    db.session.commit()
                    total_registros_guardados += len(nuevos_para_insertar)
                    logger.info(f"LOTE {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
                    logger.info(f"TOTAL ACUMULADO GUARDADOS: {total_registros_guardados}")
                    logger.info(f"TOTAL ACUMULADO DUPLICADOS EN DB IGNORADOS: {duplicados_en_db_contados}")
                else:
                    logger.info(f"LOTE {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
                
                db.session.expunge_all() 
                del registros_para_lote_db 
                registros_para_lote_db = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del LOTE {chunk_counter} liberada. ---")

        # --- Fin del bucle de filas ---

        # Procesar y guardar los registros restantes (el último lote, si lo hay)
        if registros_para_lote_db:
            chunk_counter += 1
            logger.info(f"--- Procesando LOTE FINAL {chunk_counter} ({len(registros_para_lote_db)} candidatos restantes) ---")
            
            hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
            existentes_en_db_lote = set(
                r[0] for r in db.session.query(Comentarios2025.hash_id)
                .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                .all()
            )
            nuevos_para_insertar = [
                obj for obj in registros_para_lote_db 
                if obj.hash_id not in existentes_en_db_lote
            ]
            duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

            if nuevos_para_insertar:
                db.session.bulk_save_objects(nuevos_para_insertar)
                db.session.commit()
                total_registros_guardados += len(nuevos_para_insertar)
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
            else:
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
            
            db.session.expunge_all()
            del registros_para_lote_db

        end_time = datetime.now() 
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("========== PROCESO DE CARGA COMENTARIOS 2025 FINALIZADO ==========")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS LEÍDAS DEL EXCEL: {total_filas_excel_leidas}")
        logger.info(f"TOTAL DE REGISTROS NUEVOS GUARDADOS EN LA DB: {total_registros_guardados}")
        logger.info(f"TOTAL DE DUPLICADOS EN EL ARCHIVO IGNORADOS: {duplicados_en_archivo_contados}")
        logger.info(f"TOTAL DE DUPLICADOS EN LA DB IGNORADOS: {duplicados_en_db_contados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios nuevos en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados': total_registros_guardados,
                'duplicados_ignorados_en_archivo': duplicados_en_archivo_contados,
                'duplicados_ignorados_en_db': duplicados_en_db_contados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() 
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados_hasta_error': total_registros_guardados # Registros guardados antes del error
            }
        }), 500
    
@data_mentor_bp.route('/cargar_base_loop', methods=['POST'])
def cargar_base_loop():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    try:
        # Detectar delimitador (coma o punto y coma)
        sample = archivo.read(2048).decode('utf-8')
        archivo.seek(0)
        delimiter = csv.Sniffer().sniff(sample).delimiter

        df = pd.read_csv(archivo, sep=delimiter, encoding='utf-8', on_bad_lines='skip')
    except Exception as e:
        return jsonify({"error": f"Error al leer el archivo CSV: {e}"}), 400

    try:
        columnas_db = {col.name: col.key for col in BaseLoopEstaciones.__table__.columns}

        registros = []
        for _, fila in df.iterrows():
            datos_instancia = {}
            for nombre_col_csv, valor in fila.items():
                if nombre_col_csv in columnas_db:
                    campo_para_modelo = columnas_db[nombre_col_csv]
                    datos_instancia[campo_para_modelo] = valor

            # Instanciamos así para evitar el error de keyword inválido
            instancia = BaseLoopEstaciones()
            for key, val in datos_instancia.items():
                setattr(instancia, key, val)
            registros.append(instancia)

        db.session.bulk_save_objects(registros)
        db.session.commit()

        return jsonify({"mensaje": f"{len(registros)} registros insertados correctamente"}), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        return jsonify({"error": f"Error al insertar en la base de datos: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Error al procesar el archivo: {e}"}), 500
    
@data_mentor_bp.route('/cargar_fichas_google_competencia', methods=['POST'])
def cargar_fichas_google_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            id_loop = str(fila.get('idLoop', '')).strip()
            total_review_count = str(fila.get('totalReviewCount', '')).strip()
            average_rating = str(fila.get('averageRating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not id_loop or not total_review_count or not average_rating:
                continue

            hash_id = FichasGoogleCompetencia.generar_hash(id_loop, total_review_count, average_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogleCompetencia(
                id_loop=id_loop,
                total_review_count=total_review_count,
                average_rating=average_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogleCompetencia.hash_id)
            .filter(FichasGoogleCompetencia.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500

@data_mentor_bp.route('/cargar_fichas_google', methods=['POST'])
def cargar_fichas_google():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            store_code = str(fila.get('Store Code', '')).strip()
            cantidad_de_calificaciones = str(fila.get('Cantidad de calificaciones', '')).strip()
            start_rating = str(fila.get('Star Rating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not store_code or not cantidad_de_calificaciones or not start_rating:
                continue

            hash_id = FichasGoogle.generar_hash(store_code, cantidad_de_calificaciones, start_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogle(
                store_code=store_code,
                cantidad_de_calificaciones=cantidad_de_calificaciones,
                start_rating=start_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogle.hash_id)
            .filter(FichasGoogle.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_salesforce', methods=['POST'])
def cargar_salesforce():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            valores = [
                fila.get('Estacion de Servicio: Zona', ''),
                fila.get('Número del caso', ''),
                fila.get('Estado', ''),
                fila.get('Tipificación Caso', ''),
                fila.get('Asunto', ''),
                fila.get('Fecha/Hora de apertura', ''),
                fila.get('Cantidad de Reclamos', ''),
                fila.get('Defensa al Consumidor', ''),
                fila.get('GGRR/COLA Asignado', ''),
                fila.get('Propietario del caso: Nombre completo', ''),
                fila.get('Descripción', ''),
                fila.get('Nombre del contacto: Nombre completo', ''),
                fila.get('Comentarios', ''),
                fila.get('Estacion de Servicio: Razón Social', ''),
                fila.get('Estacion de Servicio: Red', ''),
                fila.get('Estacion de Servicio: Regional', ''),
            ]

            hash_id = SalesForce.generar_hash(*valores)

            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            registro = SalesForce(
                estacion_servicio_zona=valores[0],
                numero_de_caso=valores[1],
                estado=valores[2],
                tipificacion_caso=valores[3],
                asunto=valores[4],
                fecha_apertura=valores[5],
                cantidad_reclamos=valores[6],
                defensa_consumidor=valores[7],
                ggrr_cola_asignado=valores[8],
                propietario_nombre=valores[9],
                descripcion=valores[10],
                contacto_nombre=valores[11],
                comentarios=valores[12],
                razon_social=valores[13],
                red=valores[14],
                regional=valores[15],
                hash_id=hash_id
            )

            candidatos.append(registro)
            hash_ids.append(hash_id)

        existentes = set(
            r[0] for r in db.session.query(SalesForce.hash_id)
            .filter(SalesForce.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [r for r in candidatos if r.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} casos nuevos',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_competencia', methods=['POST'])
def cargar_comentarios_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    try:
        df = pd.read_excel(archivo)

        # Normalizar la fecha
        if 'FECHA' in df.columns:
            df['FECHA'] = df['FECHA'].astype(str)


        # Renombrar columna ID si existe
        if 'ID' in df.columns:
            df.rename(columns={'ID': 'ID_ORIGINAL'}, inplace=True)

        # Reemplazamos espacios y ponemos mayúsculas para asegurar
        df.columns = [col.upper().replace(" ", "_") for col in df.columns]

        nuevos = []
        for _, fila in df.iterrows():
            hash_id = ComentariosCompetencia.generar_hash(
                fila.get('ID_ORIGINAL', ''),
                fila.get('FECHA', ''),
                fila.get('IDLOOP', ''),
                fila.get('COMENTARIO', ''),
                fila.get('RATING', ''),
                fila.get('SENTIMIENTO', ''),
                fila.get('TÓPICO', '')
            )

            # Chequeamos si ya existe
            if not ComentariosCompetencia.query.filter_by(hash_id=hash_id).first():
                nuevo = ComentariosCompetencia(
                    id_original=fila.get('ID_ORIGINAL'),
                    fecha=fila.get('FECHA'),
                    id_loop=fila.get('IDLOOP'),
                    comentario=fila.get('COMENTARIO'),
                    rating=fila.get('RATING'),
                    sentimiento=fila.get('SENTIMIENTO'),
                    topico=fila.get('TÓPICO'),
                    hash_id=hash_id
                )
                nuevos.append(nuevo)

        db.session.bulk_save_objects(nuevos)
        db.session.commit()

        return jsonify({'guardados': len(nuevos)})

    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500
    

executor = ThreadPoolExecutor(max_workers=5)


# --- Funciones auxiliares (las mismas que ya tienes) ---
def _wait_for_file_processing(file_id: str):
    """Espera a que un archivo en OpenAI cambie su estado a 'processed'."""
    start_time = time.time()
    max_wait_time = 300 
    while True:
        try:
            file_obj = client.files.retrieve(file_id=file_id)
            if file_obj.status == "processed":
                logger.info(f"Archivo {file_id} procesado por OpenAI en {time.time() - start_time:.2f} segundos.")
                return file_obj
            elif file_obj.status == "failed":
                error_message = f"El procesamiento del archivo {file_id} falló en OpenAI. Detalles: {file_obj.last_error.message if file_obj.last_error else 'Desconocido'}."
                logger.error(error_message)
                raise Exception(error_message) 
            else:
                current_wait_time = time.time() - start_time
                if current_wait_time > max_wait_time:
                    error_message = f"Tiempo de espera excedido ({max_wait_time}s) para el procesamiento del archivo {file_id}. Estado actual: {file_obj.status}"
                    logger.error(error_message)
                    raise TimeoutError(error_message) 
                logger.info(f"Archivo {file_id} en estado '{file_obj.status}', esperando... ({current_wait_time:.0f}s / {max_wait_time}s)")
                time.sleep(5)
        except Exception as e:
            logger.error(f"Error al verificar el estado del archivo {file_id}: {e}", exc_info=True)
            raise 

def _wait_for_vector_store_completion(vector_store_id: str):
    """Espera a que un Vector Store termine de asimilar todos sus archivos."""
    start_time = time.time()
    max_wait_time = 600
    while True:
        try:
            vector_store = client.beta.vector_stores.retrieve(vector_store_id)
            if vector_store.status == "completed":
                logger.info(f"Vector Store {vector_store_id} completado en {time.time() - start_time:.2f} segundos.")
                return vector_store
            elif vector_store.status in ["failed", "cancelled"]:
                error_message = f"La creación del Vector Store {vector_store_id} falló. Estado: {vector_store.status}."
                logger.error(error_message)
                raise Exception(error_message)
            else:
                current_wait_time = time.time() - start_time
                if current_wait_time > max_wait_time:
                    error_message = f"Tiempo de espera excedido ({max_wait_time}s) para la creación del Vector Store {vector_store_id}."
                    logger.error(error_message)
                    raise TimeoutError(error_message)
                logger.info(f"Vector Store {vector_store_id} en estado '{vector_store.status}', esperando... ({current_wait_time:.0f}s / {max_wait_time}s)")
                time.sleep(10)
        except Exception as e:
            logger.error(f"Error al verificar el estado del Vector Store {vector_store_id}: {e}", exc_info=True)
            raise

def _manage_vector_store_async(new_file_ids: list, old_file_id_string: str = None):
    # La lógica de esta función se integrará de forma síncrona en la nueva ruta.
    # Así que se puede dejar vacía o simplemente no usarla directamente.
    pass

# -------------------------------------------------------------------------
# RUTA PRINCIPAL PARA MÚLTIPLES TABLAS Y SUBDIVISIÓN
# -------------------------------------------------------------------------
@data_mentor_bp.route("/actualizar-conocimiento-completo-subdividido", methods=["POST"])
def actualizar_conocimiento_completo_subdividido():
    start_time = datetime.now()
    temp_dir_main = 'temp_master_json'
    temp_dir_chunks = 'temp_openai_chunks'
    
    uploaded_file_ids = []
    vector_store_id = None
    
    MAX_CHUNK_SIZE_MB = 10
    DB_QUERY_BATCH_SIZE = 10000 

    # Lista de modelos y sus nombres de sección
    TABLES_TO_INCLUDE = [
        (Comentarios2025, "comentarios_2025"),
        (Usuarios_Por_Asignacion, "usuarios_por_asignacion"),
        (DetalleApies, "detalle_apies"),
        # ¡Añade aquí el resto de tus tablas!
        # (FichasGoogle, "fichas_google"),
        # (FichasGoogleCompetencia, "fichas_google_competencia"),
        # etc...
    ]
    
    logger.info("======================================================")
    logger.info("====== INICIANDO PROCESO DE ACTUALIZACIÓN DE CONOCIMIENTO COMPLETO ======")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Tamaño máximo por archivo para OpenAI: {MAX_CHUNK_SIZE_MB} MB")
    
    try:
        if not os.path.exists(temp_dir_main): os.makedirs(temp_dir_main)
        if not os.path.exists(temp_dir_chunks): os.makedirs(temp_dir_chunks)
        master_json_path = os.path.join(temp_dir_main, "master_knowledge.json")
        
        # Paso 1: Generar el JSON maestro de forma eficiente
        logger.info("Paso 1: Generando el archivo JSON maestro de forma incremental...")
        resumen_conteos_totales = {}
        with open(master_json_path, "w", encoding="utf-8") as tmpfile:
            tmpfile.write('{\n')
            
            logger.info("Recopilando conteos totales...")
            for Model, section_name in TABLES_TO_INCLUDE:
                count = db.session.query(Model).count()
                resumen_conteos_totales[section_name] = count
                logger.info(f"  - {section_name}: {count} registros.")
            
            conteos_json_str = json.dumps(resumen_conteos_totales, indent=2, ensure_ascii=False)
            conteos_json_str_formatted = conteos_json_str.replace("{\n ", "{\n    ").replace("\n}", "\n  }")
            tmpfile.write(f'  "resumen_conteos_totales": {conteos_json_str_formatted},\n')
            tmpfile.write('  "descripcion_contenido_archivo": "Este archivo JSON contiene datos operativos y de experiencia del cliente de YPF, organizados por sección. Cada sección incluye conteos y datos detallados.",\n')

            first_section = True
            for Model, section_name in TABLES_TO_INCLUDE:
                if not first_section:
                    tmpfile.write(',\n')
                first_section = False

                logger.info(f"  - Procesando sección '{section_name}'...")
                tmpfile.write(f'  "{section_name}": [\n')

                record_count_in_section = 0
                is_first_record_in_section = True

                offset = 0
                while True:
                    batch_of_records = db.session.query(Model).limit(DB_QUERY_BATCH_SIZE).offset(offset).all()
                    if not batch_of_records:
                        break
                    
                    for item in batch_of_records:
                        if not is_first_record_in_section:
                            tmpfile.write(',\n')
                        else:
                            is_first_record_in_section = False
                        
                        json.dump(item.serialize(), tmpfile, indent=4, ensure_ascii=False)
                        record_count_in_section += 1
                    
                    offset += DB_QUERY_BATCH_SIZE
                    db.session.remove()
                    logger.info(f"    - '{section_name}': {record_count_in_section} registros procesados.")

                tmpfile.write('\n  ]') # Cierre del array de datos
            
            tmpfile.write('\n}') # Cierre del JSON maestro
            tmpfile.flush()
            tmpfile.close()
            
        logger.info(f"Archivo JSON maestro creado en: {master_json_path}. Tamaño: {os.path.getsize(master_json_path) / (1024*1024):.2f} MB")

        # Paso 2: Subdividir el JSON maestro de forma segura (CORREGIDO)
        logger.info(f"Paso 2: Subdividiendo el archivo maestro en chunks de {MAX_CHUNK_SIZE_MB} MB...")
        chunk_number = 1
        with open(master_json_path, 'r', encoding='utf-8') as master_file:
            while True:
                chunk_text = master_file.read(MAX_CHUNK_SIZE_MB * 1024 * 1024)
                if not chunk_text:
                    break
                
                # Encontrar el último carácter '}' para no cortar el JSON en un punto inválido
                last_brace_index = chunk_text.rfind('}')
                
                # NO ES NECESARIO CERRAR EL ARCHIVO MAESTRO AQUÍ
                # Con la siguiente lógica evitamos el cierre manual.
                
                # Si el último '}' no se encuentra o no es el último caracter del chunk,
                # significa que el JSON está cortado. Retrocedemos el puntero del archivo
                if last_brace_index != len(chunk_text) - 1:
                    master_file.seek(master_file.tell() - (len(chunk_text) - last_brace_index - 1))
                    chunk_text = chunk_text[:last_brace_index + 1]

                # Escribir el chunk en un archivo temporal y cerrarlo
                current_chunk_path = os.path.join(temp_dir_chunks, f"knowledge_chunk_{chunk_number}.json")
                with open(current_chunk_path, 'w', encoding='utf-8') as chunk_file:
                    chunk_file.write(chunk_text)
                
                chunk_number += 1
                
        generated_chunks = [os.path.join(temp_dir_chunks, f) for f in os.listdir(temp_dir_chunks) if f.endswith('.json')]
        if not generated_chunks:
            raise Exception("La subdivisión de archivos falló.")
        
        # Paso 3: Subir todos los chunks a OpenAI y esperar a que sean procesados
        logger.info("Paso 3: Subiendo y esperando a que todos los chunks sean procesados...")
        
        for file_path in generated_chunks:
            with open(file_path, "rb") as file_to_upload:
                uploaded_file = client.files.create(
                    file=file_to_upload,
                    purpose="assistants"
                )
            uploaded_file_ids.append(uploaded_file.id)
            logger.info(f"Chunk subido: '{os.path.basename(file_path)}', File ID: {uploaded_file.id}")
            _wait_for_file_processing(uploaded_file.id)

        # Paso 4: Crear el Vector Store con todos los archivos procesados
        logger.info("Paso 4: Todos los chunks procesados. Creando un único Vector Store...")
        
        vector_store = client.beta.vector_stores.create(
            name="Conocimiento Completo Base",
            file_ids=uploaded_file_ids
        )
        vector_store_id = vector_store.id
        logger.info(f"Vector Store creado. ID: {vector_store_id}. Esperando a que se complete la asimilación...")
        _wait_for_vector_store_completion(vector_store_id)
        
        # Paso 5: Actualizar la base de datos
        logger.info(f"Paso 5: Actualizando la base de datos...")
        existing_file_record = FileDailyID.query.first()
        old_file_id = existing_file_record.current_file_id if existing_file_record else None
        
        # Eliminar archivos obsoletos online por sus files id
        if old_file_id:
            old_file_ids_list = old_file_id.split(',')
            for old_id in old_file_ids_list:
                try:
                    # Desasociar del Vector Store
                    client.beta.vector_stores.files.delete(
                        vector_store_id=existing_file_record.current_vector_store_id,
                        file_id=old_id
                    )
                    # Eliminar el archivo del almacenamiento
                    client.files.delete(old_id)
                    logger.info(f"Archivo obsoleto '{old_id}' eliminado.")
                except Exception as e:
                    logger.warning(f"Error al eliminar archivo obsoleto '{old_id}': {e}")

        if existing_file_record:
            existing_file_record.current_file_id = ','.join(uploaded_file_ids)
            existing_file_record.current_vector_store_id = vector_store_id
        else:
            new_record = FileDailyID(
                current_file_id=','.join(uploaded_file_ids),
                current_vector_store_id=vector_store_id
            )
            db.session.add(new_record)
            
        db.session.commit()
        logger.info(f"DB actualizada. Vector Store ID: {vector_store_id}.")

        end_time = datetime.now()
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))
        
        logger.info("======================================================")
        logger.info("============ PROCESO FINALIZADO CON ÉXITO ============")
        logger.info(f"Tiempo total: {time_format}")
        logger.info("======================================================")

        return jsonify({
            "success": True,
            "message": f"Se generó el conocimiento completo a partir de {len(TABLES_TO_INCLUDE)} tablas.",
            "vector_store_id": vector_store_id,
            "archivos_creados": uploaded_file_ids,
            "tiempo_total": time_format
        }), 200

    except Exception as e:
        db.session.rollback() 
        logger.error(f"Error fatal durante el proceso: {str(e)}", exc_info=True)
        end_time = datetime.now()
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))
        return jsonify({"error": str(e), "status": 500, "tiempo_transcurrido_hasta_error": time_format_on_error}), 500
    finally:
        if os.path.exists(temp_dir_main): shutil.rmtree(temp_dir_main)
        if os.path.exists(temp_dir_chunks): shutil.rmtree(temp_dir_chunks)
        # Asegúrate de importar shutil
        # import shutil
        logger.info("Directorios temporales eliminados.")